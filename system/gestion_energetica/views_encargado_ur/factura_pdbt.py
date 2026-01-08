from datetime import datetime

from decimal import Decimal

from mongoengine.errors import DoesNotExist

from django.contrib import messages
from django.shortcuts import render, redirect
from django.http import FileResponse, Http404, HttpResponse
from django.views.decorators.cache import never_cache

from system.models import FacturaPdbt, Subestacion
from system.views import get_user
from system.decorators import login_required_custom

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


# Registro de facturas PDBT
@never_cache
@login_required_custom
def registrar_factura_pdbt(request):
    """ 
    Registra una nueva factura PDBT.

    - Permite al usuario encargado de la unidad responsable registrar una factura PDBT.
    - Verifica que el usuario tenga permisos y que la tarifa sea PDBT.
    """

    user = get_user(request)
    if not user:
        messages.error(request, "Sesión expirada.")
        return redirect('login')

    subestaciones = Subestacion.objects(unidad_responsable=user.unidad_responsable, tarifa="PDBT")
    tarifas_disponibles = set(sub.tarifa for sub in subestaciones)

    if request.method == 'POST':
        try:
            factura = FacturaPdbt(
                tipo_tarifa=request.POST.get('tipo_tarifa'),
                subestacion=Subestacion.objects.get(id=request.POST.get('subestacion')),
                dias_periodo=int(request.POST.get('dias_periodo')),
                periodo=request.POST.get('periodo'),
                consumo=Decimal(request.POST.get('consumo')),
                cargo_energia=Decimal(request.POST.get('cargo_energia')),
                importe_demanda_maxima=Decimal(request.POST.get('importe_demanda_maxima')),
                dap=Decimal(request.POST.get('dap')),
                iva=Decimal(request.POST.get('iva')),
                total_a_pagar=Decimal(request.POST.get('total_a_pagar')),
                archivo_pdf=request.FILES['archivo_pdf'],
                creado_por=user,
                status=request.POST.get('status') or "No pagada"  # Agregado
            )
            factura.save()
            messages.success(request, "Factura PDBT registrada correctamente.")
            return redirect('listar_facturas_pdbt')
        except Exception as e:
            messages.error(request, f"Error al registrar la factura: {e}")

    return render(request, 'Encargado_UR/FacturaPDBT/add_form.html', {
        'subestaciones': subestaciones,
        'tarifas_disponibles': tarifas_disponibles
    })

@never_cache
@login_required_custom
def listar_facturas_pdbt(request):
    """ 
    Listar facturas PDBT.

    - Permite al usuario encargado de la unidad responsable listar las facturas PDBT.
    - Incluye filtros por subestación y año.
    """

    user = get_user(request)
    if not user:
        messages.error(request, "Sesión expirada.")
        return redirect('login')

    subestaciones = Subestacion.objects(unidad_responsable=user.unidad_responsable, tarifa="PDBT")
    facturas = FacturaPdbt.objects(subestacion__in=subestaciones)
    tarifas_disponibles = set(sub.tarifa for sub in subestaciones)

    # Filtros
    subestacion_id = request.GET.get('subestacion')
    anio = request.GET.get('anio')

    if subestacion_id:
        facturas = facturas.filter(subestacion=subestacion_id)

    if anio:
        facturas = [f for f in facturas if f.fecha_registro.year == int(anio)]

    anios_disponibles = sorted(set(f.fecha_registro.year for f in facturas), reverse=True)

    return render(request, 'Encargado_UR/FacturaPDBT/listar_factura.html', {
        'facturas': facturas,
        'subestaciones': subestaciones,
        'anios_disponibles': anios_disponibles,
        'tarifas_disponibles': tarifas_disponibles,
        'filtros': {
            'subestacion_id': subestacion_id,
            'anio': anio
        }
    })

@never_cache
@login_required_custom
def editar_factura_pdbt(request, factura_id):
    """
    Editar una factura PDBT.   

    - Permite al usuario encargado de la unidad responsable editar una factura PDBT existente.
    - Verifica que el usuario tenga permisos y que la tarifa sea PDBT.
    """

    user = get_user(request)
    if not user:
        messages.error(request, "Sesión expirada.")
        return redirect('login')

    from mongoengine.errors import DoesNotExist
    try:
        factura = FacturaPdbt.objects.get(id=factura_id)
    except DoesNotExist:
        messages.error(request, "Factura no encontrada.")
        return redirect('listar_facturas_pdbt')

    if factura.subestacion.unidad_responsable != user.unidad_responsable:
        messages.error(request, "No tienes permiso para editar esta factura.")
        return redirect('listar_facturas_pdbt')

    subestaciones = Subestacion.objects(unidad_responsable=user.unidad_responsable, tarifa="PDBT")
    tarifas_disponibles = set(sub.tarifa for sub in subestaciones)

    if request.method == 'POST':
        try:
            factura.tipo_tarifa = request.POST.get('tipo_tarifa')
            factura.subestacion = Subestacion.objects.get(id=request.POST.get('subestacion'))
            factura.dias_periodo = int(request.POST.get('dias_periodo'))
            factura.periodo = request.POST.get('periodo')
            factura.consumo = Decimal(request.POST.get('consumo'))
            factura.cargo_energia = Decimal(request.POST.get('cargo_energia'))
            factura.importe_demanda_maxima = Decimal(request.POST.get('importe_demanda_maxima'))
            factura.dap = Decimal(request.POST.get('dap'))
            factura.iva = Decimal(request.POST.get('iva'))
            factura.total_a_pagar = Decimal(request.POST.get('total_a_pagar'))

            # Aquí se actualiza el estatus
            factura.status = request.POST.get('status')

            if 'archivo_pdf' in request.FILES:
                factura.archivo_pdf.replace(request.FILES['archivo_pdf'])

            factura.actualizado_por = user
            factura.ultima_actualizacion = datetime.now()
            factura.save()
            messages.success(request, "Factura PDBT actualizada correctamente.")
            return redirect('listar_facturas_pdbt')
        except Exception as e:
            messages.error(request, f"Error al actualizar: {e}")

    return render(request, 'Encargado_UR/FacturaPDBT/edit_form.html', {
        'factura': factura,
        'subestaciones': subestaciones,
        'tarifas_disponibles': tarifas_disponibles
    })

@never_cache
@login_required_custom
def eliminar_factura_pdbt(request, factura_id):
    """ 
    Eliminar una factura PDBT.

    - Permite al usuario encargado de la unidad responsable eliminar una factura PDBT.
    - Verifica que el usuario tenga permisos para eliminar la factura.
    """

    user = get_user(request)
    if not user:
        messages.error(request, "Sesión expirada.")
        return redirect('login')

    try:
        factura = FacturaPdbt.objects.get(id=factura_id)

        # Puedes agregar una verificación para evitar que otros usuarios la borren
        if factura.creado_por != user and not user.es_admin:
            messages.error(request, "No tienes permiso para eliminar esta factura.")
            return redirect('listar_facturas_pdbt')

        factura.delete()
        messages.success(request, "Factura eliminada correctamente.")
    except FacturaPdbt.DoesNotExist:
        messages.error(request, "La factura no existe.")
    except Exception as e:
        messages.error(request, f"Ocurrió un error al eliminar: {e}")

    return redirect('listar_facturas_pdbt')

# Descarga de PDF PDBT
@never_cache
@login_required_custom
def descargar_pdf_factura_pdbt(request, factura_id):
    """ 
    Descarga el archivo PDF de una factura PDBT.

    - Permite al usuario encargado de la unidad responsable descargar el PDF de una factura PDBT.
    - Verifica que la factura exista y tenga un archivo PDF asociado.
    """

    try:
        factura = FacturaPdbt.objects.get(id=factura_id)
        if not factura.archivo_pdf:
            raise Http404("No se encontró el archivo PDF.")
        # return FileResponse(factura.archivo_pdf, as_attachment=True, filename="factura_pdbt.pdf")
        return FileResponse(factura.archivo_pdf, filename="factura_pdbt.pdf")
    except DoesNotExist:
        raise Http404("Factura no encontrada.")

# Descargar excel de PDBT
@never_cache
@login_required_custom
def exportar_facturas_pdbt_excel(request):
    """
    Exporta las facturas PDBT a un archivo Excel.

    - Permite al usuario encargado de la unidad responsable descargar un archivo Excel con las facturas PDBT.
    - Incluye todos los campos relevantes de las facturas.
    """

    user = get_user(request)
    if not user:
        messages.error(request, "Sesión expirada.")
        return redirect('login')

    if user.rol == 'admin':
        facturas = FacturaPdbt.objects()
    else:
        subestaciones = Subestacion.objects(unidad_responsable=user.unidad_responsable)
        facturas = FacturaPdbt.objects(subestacion__in=subestaciones)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas PDBT"

    headers = [
        "Tipo de Tarifa", "Subestación", "Días de Periodo", "Periodo",
        "Consumo (kWh)", "Cargo Energía", "Importe Demanda Máxima", "DAP",
        "IVA", "Total a Pagar", "Fecha Registro", "Estado", "Creado por", "Última actualización", "Actualizado por"
    ]
    ws.append(headers)

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Estilo de encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Datos
    for factura in facturas:
        ws.append([
            factura.tipo_tarifa,
            factura.subestacion.no_servicio if factura.subestacion else '',
            factura.dias_periodo,
            factura.periodo,
            factura.consumo,
            factura.cargo_energia,
            factura.importe_demanda_maxima,
            factura.dap,
            factura.iva,
            factura.total_a_pagar,
            factura.fecha_registro.strftime("%d/%m/%Y"),
            factura.status,
            f"{factura.creado_por.nombres} {factura.creado_por.apellidos}" if factura.creado_por else '',
            factura.ultima_actualizacion.strftime("%d/%m/%Y") if factura.ultima_actualizacion else '',
            f"{factura.actualizado_por.nombres} {factura.actualizado_por.apellidos}" if factura.actualizado_por else '',
        ])

    # Estilo de celdas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # Autoajuste de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Facturas_PDBT.xlsx'
    wb.save(response)
    return response