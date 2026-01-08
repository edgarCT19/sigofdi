from datetime import datetime

from decimal import Decimal

from bson import ObjectId

from mongoengine.errors import DoesNotExist

from django.contrib import messages
from django.shortcuts import render, redirect
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST
from django.utils.encoding import smart_str
from django.core.exceptions import ValidationError

from system.models import FacturaEnergeticaTriple, Subestacion
from system.views import get_user
from system.decorators import login_required_custom

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill



# Registro de facturas triples, es decir, de la tarifa GDBT, GDMTH y GDMTO
@never_cache
@login_required_custom
def listar_facturas_triple(request):
    """
    Listar las facturas de tarifa triple (GDBT, GDMTH, GDMTO).

    - Permite al usuario encargado de la unidad responsable ver las facturas registradas    
    - Filtra las facturas por subestación, tipo de tarifa y año.
    - Muestra las subestaciones disponibles y las tarifas asociadas.
    """

    user = get_user(request)
    if not user:
        messages.error(request, "Sesión expirada.")
        return redirect('login')
  
    # Subestaciones de la UR del usuario
    subestaciones = Subestacion.objects(unidad_responsable=user.unidad_responsable)

    # Extraer tarifas disponibles según sus subestaciones
    tarifas_disponibles = set(sub.tarifa for sub in subestaciones)

    # Obtener todas las facturas ligadas a esas subestaciones
    facturas = FacturaEnergeticaTriple.objects(subestacion__in=subestaciones)

    # Filtros
    subestacion_id = request.GET.get('subestacion')
    tipo_tarifa = request.GET.get('tipo_tarifa')
    anio = request.GET.get('anio')

    if subestacion_id:
        facturas = facturas.filter(subestacion=subestacion_id)

    if tipo_tarifa:
        facturas = facturas.filter(tipo_tarifa=tipo_tarifa)

    if anio:
        facturas = [f for f in facturas if f.fecha_registro.year == int(anio)]

    # Obtener años únicos para el filtro por año
    anios_disponibles = sorted(
        set(f.fecha_registro.year for f in FacturaEnergeticaTriple.objects(subestacion__in=subestaciones)),
        reverse=True
    )

    return render(request, 'Encargado_UR/Facturas/facturas.html', {
        'facturas': facturas,
        'subestaciones': subestaciones,
        'anios_disponibles': anios_disponibles,
        'tarifas_disponibles': tarifas_disponibles,  # Se envía al template
        'filtros': {
            'subestacion_id': subestacion_id,
            'tipo_tarifa': tipo_tarifa,
            'anio': anio
        }
    })

@never_cache
@login_required_custom
def registrar_factura_triple(request):
    """
    Registrar una nueva factura de tarifa triple (GDBT, GDMTH, GDMTO).   

    - Permite al usuario encargado de la unidad responsable registrar una factura.
    - Verifica que el usuario tenga acceso a las subestaciones con tarifas GDMTH, GDMTO o GDBT.
    - Muestra un formulario para ingresar los datos de la factura.
    """

    user = get_user(request)
    if not user:
        messages.error(request, "Sesión expirada.")
        return redirect('login')

    subestaciones = Subestacion.objects(unidad_responsable=user.unidad_responsable, tarifa__in=['GDMTH', 'GDMTO', 'GDBT'])
    tarifas_disponibles = set(sub.tarifa for sub in subestaciones)

    if not subestaciones:
        messages.error(request, "No tienes subestaciones con tarifa GDMTH, GDMTO o GDBT.")
        return redirect('listar_facturas_triple')

    if request.method == 'POST':
        try:
            factura = FacturaEnergeticaTriple(
                tipo_tarifa=request.POST.get('tipo_tarifa'),
                subestacion=Subestacion.objects.get(id=request.POST.get('subestacion')),
                dias_periodo=int(request.POST.get('dias_periodo')),
                periodo=request.POST.get('periodo'),
                consumo=Decimal(request.POST.get('consumo')),
                demanda_maxima=int(request.POST.get('demanda_maxima')),
                factor_potencia=Decimal(request.POST.get('factor_potencia')),
                factor_carga=int(request.POST.get('factor_carga')),
                cargo_energia=Decimal(request.POST.get('cargo_energia')),
                importe_demanda_maxima=Decimal(request.POST.get('importe_demanda_maxima')),
                importe_bt=Decimal(request.POST.get('importe_bt')),
                importe_fp=Decimal(request.POST.get('importe_fp')),
                dap=Decimal(request.POST.get('dap')),
                iva=Decimal(request.POST.get('iva')),
                total_a_pagar=Decimal(request.POST.get('total_a_pagar')),
                archivo_pdf=request.FILES['archivo_pdf'],
                fecha_vencimiento=request.POST.get('fecha_vencimiento'),
                status=request.POST.get('status'),
                creado_por=user
            )
            factura.save()
            messages.success(request, "Factura registrada correctamente.")
            return redirect('listar_facturas_triple')
        except Exception as e:
            messages.error(request, f"Ocurrió un error: {e}")

    return render(request, 'Encargado_UR/Facturas/add_form.html', {
        'subestaciones': subestaciones,
        'tarifas_disponibles': tarifas_disponibles
    })
    
@never_cache
@login_required_custom
def editar_factura_triple(request, factura_id):
    """ 
    Editar una factura de tarifa triple (GDBT, GDMTH, GDMTO).

    - Permite al usuario encargado de la unidad responsable editar los datos de una factura existente.
    - Verifica que el usuario tenga acceso a la factura y a la subestación asociada.
    - Muestra un formulario con los datos actuales de la factura para su edición.
    """

    user = get_user(request)
    if not user:
        messages.error(request, "Sesión expirada.")
        return redirect('login')

    from mongoengine.errors import DoesNotExist

    try:
        factura = FacturaEnergeticaTriple.objects.get(id=factura_id)
    except DoesNotExist:
        messages.error(request, "Factura no encontrada.")
        return redirect('listar_facturas_triple')

    if factura.subestacion.unidad_responsable != user.unidad_responsable:
        messages.error(request, "No tienes permiso para editar esta factura.")
        return redirect('listar_facturas_triple')

    subestaciones = Subestacion.objects(unidad_responsable=user.unidad_responsable, tarifa__in=['GDMTH', 'GDMTO', 'GDBT'])
    tarifas_disponibles = set(sub.tarifa for sub in subestaciones)

    if request.method == 'POST':
        try:
            factura.tipo_tarifa = request.POST.get('tipo_tarifa')
            factura.subestacion = Subestacion.objects.get(id=request.POST.get('subestacion'))
            factura.dias_periodo = int(request.POST.get('dias_periodo'))
            factura.periodo = request.POST.get('periodo')
            factura.consumo = Decimal(request.POST.get('consumo'))
            factura.demanda_maxima = int(request.POST.get('demanda_maxima'))
            factura.factor_potencia = Decimal(request.POST.get('factor_potencia'))
            factura.factor_carga = int(request.POST.get('factor_carga'))
            factura.cargo_energia = Decimal(request.POST.get('cargo_energia'))
            factura.importe_demanda_maxima = Decimal(request.POST.get('importe_demanda_maxima'))
            factura.importe_bt = Decimal(request.POST.get('importe_bt'))
            factura.importe_fp = Decimal(request.POST.get('importe_fp'))
            factura.dap = Decimal(request.POST.get('dap'))
            factura.iva = Decimal(request.POST.get('iva'))
            factura.total_a_pagar = Decimal(request.POST.get('total_a_pagar'))
            factura.status = request.POST.get('status')
            factura.fecha_vencimiento = request.POST.get('fecha_vencimiento')

            if 'archivo_pdf' in request.FILES:
                factura.archivo_pdf.replace(request.FILES['archivo_pdf'])

            factura.actualizado_por = user
            factura.ultima_actualizacion = datetime.now()
            factura.save()
            messages.success(request, "Factura actualizada correctamente.")
            return redirect('listar_facturas_triple')
        except Exception as e:
            messages.error(request, f"Ocurrió un error: {e}")

    tipos_tarifa = ['GDMTH', 'GDMTO', 'GDBT']

    return render(request, 'Encargado_UR/Facturas/edit_form.html', {
        'factura': factura,
        'subestaciones': subestaciones,
        'tipos_tarifa': tipos_tarifa,
        'tarifas_disponibles': tarifas_disponibles
    })

@require_POST
@login_required_custom
def eliminar_factura_triple(request, factura_id):  
    """
    Eliminar una factura de tarifa triple (GDBT, GDMTH, GDMTO).   

    - Permite al usuario encargado de la unidad responsable eliminar una factura existente.
    - Verifica que el usuario tenga permisos para eliminar la factura.
    - Elimina el archivo PDF asociado si existe.
    """

    user = get_user(request)

    if not factura_id or not ObjectId.is_valid(factura_id):
        return JsonResponse({"success": False, "message": "ID de factura inválido."}, status=400)

    try:
        factura = FacturaEnergeticaTriple.objects.get(id=factura_id)

        # Validar permisos
        if factura.creado_por and factura.creado_por != user:
            return JsonResponse({"success": False, "message": "No tienes permiso para eliminar esta factura."}, status=403)

        if factura.subestacion.unidad_responsable != user.unidad_responsable:
            return JsonResponse({"success": False, "message": "No tienes acceso a esta subestación."}, status=403)

        # Eliminar archivo si existe
        if factura.archivo_pdf:
            factura.archivo_pdf.delete()

        factura.delete()
        return JsonResponse({"success": True, "message": "Factura eliminada correctamente."})

    except DoesNotExist:
        return JsonResponse({"success": False, "message": "La factura no existe."}, status=404)
    except ValidationError:
        return JsonResponse({"success": False, "message": "Error al procesar la solicitud."}, status=400)

# Descarga de PDF facturas triple (GDBT, GDMTH, GDMTO)
def tiene_tarifa(user, tarifa_objetivo):
    """ Verifica si el usuario tiene acceso a una tarifa específica en sus subestaciones.

    - Permite determinar si el usuario puede acceder a facturas de una tarifa específica.
    - Revisa las subestaciones asociadas al usuario y verifica si alguna tiene la tarifa objetivo.
    """

    subestaciones = Subestacion.objects(unidad_responsable=user.unidad_responsable)
    return any(sub.tarifa == tarifa_objetivo for sub in subestaciones)

def descargar_pdf_factura(request, factura_id):
    """ Descarga el archivo PDF de una factura de tarifa triple (GDBT, GDMTH, GDMTO).

    - Permite al usuario encargado de la unidad responsable descargar el PDF de una factura.
    - Verifica que la factura exista y tenga un archivo PDF asociado.
    - Si la factura no existe o no tiene un archivo PDF, devuelve un error 404.
    """

    try:
        factura = FacturaEnergeticaTriple.objects.get(id=factura_id)
        archivo = factura.archivo_pdf
        archivo.seek(0)  # Reiniciar el puntero al inicio
        return FileResponse(archivo, content_type='application/pdf', filename='factura.pdf')
    except (DoesNotExist, Exception):
        raise Http404("Archivo no encontrado.")

# Descargar reporte en excel de las facturas triple
@never_cache
@login_required_custom
def exportar_facturas_triple_excel(request):
    """ Exporta las facturas de tarifa triple (GDBT, GDMTH, GDMTO) a un archivo Excel.

    - Permite al usuario encargado de la unidad responsable descargar un archivo Excel con las facturas triple.         
    - Incluye todos los campos relevantes de las facturas.
    - Verifica que el usuario tenga acceso a las subestaciones con tarifas GDMTH, GDMTO o GDBT.
    """
    
    user = get_user(request)

    if not user:
        messages.error(request, "Sesión expirada.")
        return redirect('login')

    if user.rol == 'admin':
        ur_id = request.GET.get('ur')
        if ur_id:
            subestaciones = Subestacion.objects(unidad_responsable=ur_id)
        else:
            subestaciones = Subestacion.objects()
    else:
        subestaciones = Subestacion.objects(unidad_responsable=user.unidad_responsable)

    facturas = FacturaEnergeticaTriple.objects(subestacion__in=subestaciones)

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas Triple"

    # Encabezados
    headers = [
        "Unidad Responsable", "Subestación", "Tarifa", "Periodo", "Días", "Consumo",
        "Demanda Máxima", "Factor Potencia", "Factor Carga", "Cargo Energía",
        "Importe Demanda", "Importe BT", "Importe FP", "DAP", "IVA", "Total a Pagar",
        "Estatus", "Creado por", "Fecha Registro", "Última actualización", "Actualizado por"
    ]
    ws.append(headers)

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Aplicar estilo a encabezados
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Agregar datos
    for factura in facturas:
        ws.append([
            factura.subestacion.unidad_responsable.nombre,
            f"N° {factura.subestacion.no_servicio}",
            factura.tipo_tarifa,
            factura.periodo,
            factura.dias_periodo,
            factura.consumo,
            factura.demanda_maxima,
            factura.factor_potencia,
            factura.factor_carga,
            factura.cargo_energia,
            factura.importe_demanda_maxima,
            factura.importe_bt,
            factura.importe_fp,
            factura.dap,
            factura.iva,
            factura.total_a_pagar,
            factura.status,
            f"{factura.creado_por.nombres} {factura.creado_por.apellidos}" if factura.creado_por else '',
            factura.fecha_registro.strftime("%Y-%m-%d"),
            factura.ultima_actualizacion.strftime("%Y-%m-%d") if factura.ultima_actualizacion else '',
            f"{factura.actualizado_por.nombres} {factura.actualizado_por.apellidos}" if factura.actualizado_por else ''
        ])

    # Estilos para celdas de datos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajuste automático de columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

    # Preparar respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nombre_archivo = "Facturas_Triple.xlsx"
    response['Content-Disposition'] = f'attachment; filename={smart_str(nombre_archivo)}'
    wb.save(response)
    return response