from datetime import datetime

from bson import ObjectId
from bson.errors import InvalidId

from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.utils.cache import add_never_cache_headers
from django.views.decorators.cache import never_cache
from django.contrib import messages

from system.models import FacturaEnergeticaTriple, FacturaPdbt, Subestacion, UnidadResponsable
from system.views import get_user
from system.decorators import login_required_custom

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

# Funiones para gestionar facturas energéticas (Triple y PDBT) en el panel de administración.
@never_cache
@login_required_custom
def listar_facturas_admin(request):
    """
    Vista para listar facturas en el panel de administración.
    Permite filtrar por unidad responsable (UR), subestación, tipo de tarifa y año.
    Solo accesible para usuarios con rol 'admin'.
    """
    # Obtener usuario actual
    user = get_user(request)
    if not user or user.rol not in ["admin", "admin_energia", "admin_ambiental"]:
        return redirect('inicio')

    # Parámetros de filtrado desde la URL
    ur_id = request.GET.get('ur')               # ID de unidad responsable
    sub_id = request.GET.get('subestacion')     # ID de subestación
    tipo_tarifa = request.GET.get('tipo_tarifa') # Tarifa específica
    anio = request.GET.get('anio')              # Año de registro

    # Catálogos para desplegar en el formulario de filtros
    unidades = UnidadResponsable.objects()
    subestaciones = (
        Subestacion.objects(unidad_responsable=ur_id)
        if ur_id else Subestacion.objects()
    )

    # Obtener todas las facturas sin filtrar inicialmente
    facturas_triple = FacturaEnergeticaTriple.objects()
    facturas_pdbt = FacturaPdbt.objects()

    # Filtrar por unidad responsable
    if ur_id:
        subs_ur = Subestacion.objects(unidad_responsable=ur_id)
        facturas_triple = facturas_triple.filter(subestacion__in=subs_ur)
        facturas_pdbt = facturas_pdbt.filter(subestacion__in=subs_ur)

    # Filtrar por subestación
    if sub_id:
        facturas_triple = facturas_triple.filter(subestacion=sub_id)
        facturas_pdbt = facturas_pdbt.filter(subestacion=sub_id)

    # Filtrar por tipo de tarifa
    if tipo_tarifa:
        facturas_triple = facturas_triple.filter(tipo_tarifa=tipo_tarifa)
        facturas_pdbt = facturas_pdbt.filter(tipo_tarifa=tipo_tarifa)

    # Filtrar por año de registro
    if anio:
        facturas_triple = [
            f for f in facturas_triple if f.fecha_registro.year == int(anio)
        ]
        facturas_pdbt = [
            f for f in facturas_pdbt if f.fecha_registro.year == int(anio)
        ]

    # Obtener lista de años disponibles en ambas colecciones
    anios_disponibles = sorted(
        set(f.fecha_registro.year for f in FacturaEnergeticaTriple.objects()) |
        set(f.fecha_registro.year for f in FacturaPdbt.objects()),
        reverse=True
    )

    # Contexto para la plantilla
    context = {
        'facturas_triple': facturas_triple,
        'facturas_pdbt': facturas_pdbt,
        'unidades': unidades,
        'subestaciones': subestaciones,
        'anios_disponibles': anios_disponibles,
        'filtros': {
            'ur_id': ur_id,
            'sub_id': sub_id,
            'tipo_tarifa': tipo_tarifa,
            'anio': anio,
        }
    }

    # Renderizar plantilla con datos filtrados
    return render(request, 'systemsigo/Facturas/facturas_admin.html', context)

def api_subestaciones_por_ur(request, ur_id):
    """
    API endpoint para obtener las subestaciones asociadas a una Unidad Responsable específica.
    
    Args:
        request (HttpRequest): La solicitud HTTP recibida.
        ur_id (str): ID de la Unidad Responsable en formato string.

    Returns:
        JsonResponse: Lista de subestaciones con su información formateada o un mensaje de error.
    """
    # --- Validación del ID recibido ---
    try:
        ur_obj_id = ObjectId(ur_id)  # Convertir el ID de cadena a ObjectId
    except (InvalidId, TypeError, ValueError) as e:
        return JsonResponse({"error": "ID inválido: " + str(e)}, status=400)

    # --- Obtención de subestaciones asociadas ---
    try:
        subestaciones = Subestacion.objects(unidad_responsable=ur_obj_id)

        data = [{
            "id": str(s.id),
            "nombre": f"Medidor: {s.no_medidor} | Servicio: {s.no_servicio} | Tarifa: {s.tarifa}"
        } for s in subestaciones]

        return JsonResponse(data, safe=False)

    except Exception as e:
        return JsonResponse({"error": "Error interno: " + str(e)}, status=500)

# CRUD FACTURA TRIPLE
@never_cache
@login_required_custom
def listar_facturas_triple_admin(request):
    """
    Vista para listar facturas energéticas de tipo 'triple' en el panel de administración.
    - Solo los usuarios con rol 'admin' pueden acceder a esta vista.
    - Permite filtrar las facturas por unidad responsable, subestación, tipo de tarifa y año.
    - Muestra una lista de facturas que cumplen con los criterios de filtrado.
    """

    user = get_user(request)
    if not user or user.rol not in ["admin", "admin_energia", "admin_ambiental"]:
        return redirect('inicio')

    # Parámetros de filtrado desde la URL
    ur_id = request.GET.get('ur')
    sub_id = request.GET.get('subestacion')
    tipo_tarifa = request.GET.get('tipo_tarifa')
    anio = request.GET.get('anio')

    # Catálogos para filtros
    unidades = UnidadResponsable.objects()
    subestaciones = Subestacion.objects(unidad_responsable=ur_id) if ur_id else Subestacion.objects()

    # Obtener todas las facturas sin filtrar inicialmente
    facturas = FacturaEnergeticaTriple.objects()

    # Aplicar filtros
    if ur_id:
        subs_ur = Subestacion.objects(unidad_responsable=ur_id)
        facturas = facturas.filter(subestacion__in=subs_ur)

    if sub_id:
        facturas = facturas.filter(subestacion=sub_id)

    if tipo_tarifa:
        facturas = facturas.filter(tipo_tarifa=tipo_tarifa)

    if anio:
        facturas = [f for f in facturas if f.fecha_registro.year == int(anio)]

    context = {
        'facturas': facturas,
        'unidades': unidades,
        'subestaciones': subestaciones,
        'filtros': {
            'ur_id': ur_id,
            'sub_id': sub_id,
            'tipo_tarifa': tipo_tarifa,
            'anio': anio,
        }
    }

    return render(request, 'systemsigo/Facturas/Factura_Triple/list_triple.html', context)

@never_cache
@login_required_custom
def crear_factura_triple(request):
    """
    Vista para crear una factura energética de tipo 'triple'.
    - Los administradores pueden seleccionar cualquier Unidad Responsable (UR).
    - Los demás usuarios solo pueden registrar facturas para su propia UR.
    - Valida que la subestación seleccionada pertenezca a la UR indicada.
    - Guarda la factura con los datos proporcionados en el formulario.
    """

    user = get_user(request)

    # Usuarios administradores pueden ver todas las URs
    if user and user.rol in ["admin", "admin_energia", "admin_ambiental"]:
        urs = UnidadResponsable.objects.all()
    # Usuarios normales ven solo su propia UR
    else:
        urs = UnidadResponsable.objects(id=user.unidad_responsable.id)

    if request.method == "POST":
        try:
            ur_id = request.POST.get("unidad_responsable")
            sub_id = request.POST.get("subestacion")
            tipo_tarifa = request.POST.get("tipo_tarifa")

            consumo = request.POST.get("consumo")
            periodo = request.POST.get("periodo")
            demanda_maxima = int(request.POST.get("demanda_maxima"))
            factor_potencia = request.POST.get("factor_potencia")
            factor_carga = int(request.POST.get("factor_carga"))
            cargo_energia = request.POST.get("cargo_energia")
            importe_demanda_maxima = request.POST.get("importe_demanda_maxima")
            importe_bt = request.POST.get("importe_bt")
            importe_fp = request.POST.get("importe_fp")
            dap = request.POST.get("dap")
            iva = request.POST.get("iva")
            total_a_pagar = request.POST.get("total_a_pagar")
            archivo_pdf = request.FILES.get("archivo_pdf")
            fecha_vencimiento = request.POST.get("fecha_vencimiento")
            dias_periodo = int(request.POST.get("dias_periodo"))

            ur_obj = ObjectId(ur_id)
            sub_obj = ObjectId(sub_id)

            subestacion_validada = Subestacion.objects(
                id=sub_obj,
                unidad_responsable=ur_obj
            ).first()
            if not subestacion_validada:
                raise ValueError("La subestación no corresponde a la UR seleccionada.")

            factura = FacturaEnergeticaTriple(
                tipo_tarifa=tipo_tarifa,
                subestacion=subestacion_validada,
                dias_periodo=dias_periodo,
                periodo=periodo,
                consumo=consumo,
                demanda_maxima=demanda_maxima,
                factor_potencia=factor_potencia,
                factor_carga=factor_carga,
                cargo_energia=cargo_energia,
                importe_demanda_maxima=importe_demanda_maxima,
                importe_bt=importe_bt,
                importe_fp=importe_fp,
                dap=dap,
                iva=iva,
                total_a_pagar=total_a_pagar,
                archivo_pdf=archivo_pdf,
                fecha_vencimiento=fecha_vencimiento,
                creado_por=user,
            )
            factura.save()

            messages.success(request, "Factura creada correctamente.")
            return redirect("listar_facturas_triple_admin")

        except Exception as e:
            return render(
                request,
                "systemsigo/Facturas/Factura_Triple/add_form.html",
                {
                    "error": str(e),
                    "urs": urs,
                    "user_rol": user.rol,
                }
            )

    # Si es GET, renderizar el formulario vacío
    return render(
        request,
        "systemsigo/Facturas/Factura_Triple/add_form.html",
        {
            "urs": urs,
            "user_rol": user.rol,
        }
    )
    
@never_cache
@login_required_custom
def editar_factura_triple_admin(request, f_id):
    """
    Vista para editar facturas energéticas de tipo 'triple' en el panel de administración.
    - Solo los usuarios con rol 'admin' pueden editar cualquier factura.
    - Los usuarios no administradores solo pueden editar facturas pertenecientes a su Unidad Responsable.
    - Valida la relación entre la Unidad Responsable y la subestación seleccionada.
    - Permite actualizar todos los campos relevantes de la factura, incluyendo la carga de un nuevo PDF.
    - Registra el usuario y la fecha de la última modificación.
    - Si la factura no existe o el usuario no tiene permisos, muestra un mensaje de error.
    """

    user = get_user(request)
    factura = FacturaEnergeticaTriple.objects(id=f_id).first()

    if not factura:
        return render(request, "error.html", {"mensaje": "Factura no encontrada"})

    # Validar permiso: admin puede editar todo
    if user.rol != "admin":
        if not factura.subestacion or factura.subestacion.unidad_responsable.id != user.unidad_responsable.id:
            return render(request, "error.html", {"mensaje": "No tienes permiso para editar esta factura."})

    # URs disponibles para el formulario
    urs = UnidadResponsable.objects.all() if user.rol == "admin" else UnidadResponsable.objects(id=user.unidad_responsable.id)

    if request.method == "POST":
        try:
            ur_id = request.POST.get("unidad_responsable")
            sub_id = request.POST.get("subestacion")
            tipo_tarifa = request.POST.get("tipo_tarifa")

            ur_obj = ObjectId(ur_id)
            sub_obj = ObjectId(sub_id)

            subestacion_validada = Subestacion.objects(id=sub_obj, unidad_responsable=ur_obj).first()
            if not subestacion_validada:
                raise ValueError("La subestación no corresponde a la UR seleccionada.")

            factura.tipo_tarifa = tipo_tarifa
            factura.subestacion = subestacion_validada
            factura.dias_periodo = int(request.POST.get("dias_periodo"))
            factura.periodo = request.POST.get("periodo")
            factura.consumo = request.POST.get("consumo")
            factura.demanda_maxima = int(request.POST.get("demanda_maxima"))
            factura.factor_potencia = request.POST.get("factor_potencia")
            factura.factor_carga = int(request.POST.get("factor_carga"))
            factura.cargo_energia = request.POST.get("cargo_energia")
            factura.importe_demanda_maxima = request.POST.get("importe_demanda_maxima")
            factura.importe_bt = request.POST.get("importe_bt")
            factura.importe_fp = request.POST.get("importe_fp")
            factura.dap = request.POST.get("dap")
            factura.iva = request.POST.get("iva")
            factura.total_a_pagar = request.POST.get("total_a_pagar")
            factura.status = request.POST.get("status")
            factura.fecha_vencimiento = request.POST.get("fecha_vencimiento")

            if request.FILES.get("archivo_pdf"):
                factura.archivo_pdf = request.FILES.get("archivo_pdf")

            factura.actualizado_por = user
            factura.ultima_actualizacion = datetime.now()
            factura.save()

            messages.success(request, "Información de factura actualizada correctamente.")
            return redirect("listar_facturas_triple_admin")

        except Exception as e:
            ur_id = request.POST.get("unidad_responsable")
            subestaciones = []
            if ur_id:
                try:
                    ur_obj_id = ObjectId(ur_id)
                    subestaciones = Subestacion.objects(unidad_responsable=ur_obj_id)
                except:
                    pass

            return render(request, "systemsigo/Facturas/Factura_Triple/edit_form.html", {
                "error": str(e),
                "factura": factura,
                "urs": urs,
                "user_rol": user.rol,
                "subestaciones": subestaciones,
            })

    subestaciones = Subestacion.objects(unidad_responsable=factura.subestacion.unidad_responsable.id)

    return render(request, "systemsigo/Facturas/Factura_Triple/edit_form.html", {
        "factura": factura,
        "urs": urs,
        "user_rol": user.rol,
        "subestaciones": subestaciones
    })

@never_cache
@login_required_custom
def eliminar_factura_triple_admin(request, f_id):
    """
    Vista para eliminar una factura energética de tipo 'triple' desde el panel de administración.
    - Solo los usuarios con rol 'admin' pueden eliminar facturas.
    - Recibe el ID de la factura a eliminar.
    - Devuelve una respuesta JSON indicando el éxito o fracaso de la operación.
    - Maneja errores si la factura no existe o si el usuario no tiene permisos.
    - Acepta solo peticiones POST para realizar la eliminación.
    - Si la factura no se encuentra, devuelve un error 404.
    """

    user = get_user(request)
    factura = FacturaEnergeticaTriple.objects(id=f_id).first()

    if not factura:
        return JsonResponse({'success': False, 'message': 'Factura no encontrada'}, status=404)

    if not user or user.rol not in ["admin", "admin_energia", "admin_ambiental"]:
        return JsonResponse({'success': False, 'message': 'No autorizado'}, status=403)

    factura.delete()
    return JsonResponse({'success': True})

# CRUD FACTURA PDBT
@never_cache
@login_required_custom
def listar_facturas_pdbt_admin(request):
    """
    Vista para listar facturas PDBT en el panel de administración.
    - Solo los usuarios con rol 'admin' pueden acceder a esta vista.
    - Permite filtrar las facturas por unidad responsable, subestación, tipo de tarifa y año.
    - Muestra una lista de facturas que cumplen con los criterios de filtrado.
    """

    user = get_user(request)
    if not user or user.rol not in ["admin", "admin_energia", "admin_ambiental"]:
        return redirect('inicio')

    # Parámetros de filtrado desde la URL
    ur_id = request.GET.get('ur')
    sub_id = request.GET.get('subestacion')
    tipo_tarifa = request.GET.get('tipo_tarifa')
    anio = request.GET.get('anio')

    # Catálogos para filtros
    unidades = UnidadResponsable.objects()
    subestaciones = Subestacion.objects(unidad_responsable=ur_id) if ur_id else Subestacion.objects()

    # Obtener todas las facturas sin filtrar inicialmente
    facturas = FacturaPdbt.objects()

    # Aplicar filtros
    if ur_id:
        subs_ur = Subestacion.objects(unidad_responsable=ur_id)
        facturas = facturas.filter(subestacion__in=subs_ur)

    if sub_id:
        facturas = facturas.filter(subestacion=sub_id)

    if tipo_tarifa:
        facturas = facturas.filter(tipo_tarifa=tipo_tarifa)

    if anio:
        facturas = [f for f in facturas if f.fecha_registro.year == int(anio)]

    context = {
        'facturas': facturas,
        'unidades': unidades,
        'subestaciones': subestaciones,
        'filtros': {
            'ur_id': ur_id,
            'sub_id': sub_id,
            'tipo_tarifa': tipo_tarifa,
            'anio': anio,
        }
    }

    return render(request, 'systemsigo/Facturas/Factura_PDBT/list_pdbt.html', context)

@never_cache
@login_required_custom
def editar_factura_pdbt_admin(request, f_id):
    """
    Vista para editar una factura del tipo PDBT desde el panel de administración.
    - Los usuarios con rol 'admin' pueden editar cualquier factura.
    - Los usuarios no administradores solo pueden editar facturas de su propia Unidad Responsable.
    - Valida que la subestación seleccionada pertenezca a la Unidad Responsable indicada.
    - Permite actualizar todos los campos relevantes de la factura, incluyendo la carga de un nuevo PDF.
    - Registra el usuario y la fecha de la última modificación.
    - Si la factura no existe o el usuario no tiene permisos, muestra un mensaje de error.
    """

    user = get_user(request)
    factura = FacturaPdbt.objects(id=f_id).first()

    if not factura:
        return render(request, "error.html", {"mensaje": "Factura no encontrada"})

    # Validación de permisos
    if not user or user.rol not in ["admin", "admin_energia", "admin_ambiental"]:
        if not factura.subestacion or factura.subestacion.unidad_responsable.id != user.unidad_responsable.id:
            return render(request, "error.html", {"mensaje": "No tienes permiso para editar esta factura."})

    # Unidades responsables disponibles para el usuario
    urs = UnidadResponsable.objects.all() if user.rol == "admin" else UnidadResponsable.objects(id=user.unidad_responsable.id)

    if request.method == "POST":
        try:
            ur_id = request.POST.get("unidad_responsable")
            sub_id = request.POST.get("subestacion")

            ur_obj = ObjectId(ur_id)
            sub_obj = ObjectId(sub_id)

            subestacion_validada = Subestacion.objects(id=sub_obj, unidad_responsable=ur_obj).first()
            if not subestacion_validada:
                raise ValueError("La subestación no corresponde a la UR seleccionada.")

            factura.subestacion = subestacion_validada
            factura.dias_periodo = int(request.POST.get("dias_periodo"))
            factura.periodo = request.POST.get("periodo")
            factura.consumo = request.POST.get("consumo")
            factura.cargo_energia = request.POST.get("cargo_energia")
            factura.importe_demanda_maxima = request.POST.get("importe_demanda_maxima")
            factura.dap = request.POST.get("dap")
            factura.iva = request.POST.get("iva")
            factura.total_a_pagar = request.POST.get("total_a_pagar")
            factura.fecha_vencimiento = request.POST.get("fecha_vencimiento")
            factura.status = request.POST.get("status")

            if request.FILES.get("archivo_pdf"):
                factura.archivo_pdf = request.FILES.get("archivo_pdf")

            factura.actualizado_por = user
            factura.ultima_actualizacion = datetime.now()
            factura.save()

            messages.success(request, "Información de factura actualizada correctamente.")
            return redirect("listar_facturas_pdbt_admin")

        except Exception as e:
            return render(request, "systemsigo/Facturas/Factura_PDBT/edit_form.html", {
                "error": str(e),
                "factura": factura,
                "urs": urs,
                "user_rol": user.rol,
                "subestaciones": [],
            })

    # Subestaciones según UR
    subestaciones = Subestacion.objects(unidad_responsable=factura.subestacion.unidad_responsable.id)

    return render(request, "systemsigo/Facturas/Factura_PDBT/edit_form.html", {
        "factura": factura,
        "urs": urs,
        "user_rol": user.rol,
        "subestaciones": subestaciones,
    })

@never_cache
@login_required_custom
def crear_factura_pdbt(request):
    """
    Vista para crear una nueva factura del tipo PDBT.
    - Restringida a usuarios autenticados.
    - Diferencia el acceso de datos según el rol del usuario.
    - Valida que la subestación seleccionada pertenezca a la Unidad Responsable (UR) elegida.
    """

    user = get_user(request)

    urs = UnidadResponsable.objects.all() if user.rol == "admin" else UnidadResponsable.objects(id=user.unidad_responsable.id)

    if request.method == "POST":
        try:
            ur_id = request.POST.get("unidad_responsable")
            sub_id = request.POST.get("subestacion")

            ur_obj = ObjectId(ur_id)
            sub_obj = ObjectId(sub_id)

            subestacion_validada = Subestacion.objects(id=sub_obj, unidad_responsable=ur_obj).first()
            if not subestacion_validada:
                raise ValueError("La subestación no pertenece a la UR seleccionada.")

            factura = FacturaPdbt(
                subestacion=subestacion_validada,
                dias_periodo=int(request.POST.get("dias_periodo")),
                periodo=request.POST.get("periodo"),
                consumo=request.POST.get("consumo"),
                cargo_energia=request.POST.get("cargo_energia"),
                importe_demanda_maxima=request.POST.get("importe_demanda_maxima"),
                dap=request.POST.get("dap"),
                iva=request.POST.get("iva"),
                total_a_pagar=request.POST.get("total_a_pagar"),
                archivo_pdf=request.FILES.get("archivo_pdf"),
                fecha_vencimiento=request.POST.get("fecha_vencimiento"),
                status=request.POST.get("status"),
                creado_por=user,
            )
            factura.save()
            messages.success(request, "Factura creada correctamente.")
            return redirect("listar_facturas_pdbt_admin")
        except Exception as e:
            return render(request, "systemsigo/Facturas/Factura_PDBT/add_form.html", {
                "error": str(e),
                "urs": urs,
                "user_rol": user.rol,
                "subestaciones": [],
            })

    return render(request, "systemsigo/Facturas/Factura_PDBT/add_form.html", {
        "urs": urs,
        "user_rol": user.rol,
        "subestaciones": [],
    })

@never_cache
@login_required_custom
def eliminar_factura_pdbt_admin(request, f_id):
    """
    Vista para eliminar una factura PDBT desde la interfaz de administrador.
    
    Args:
        request (HttpRequest): Objeto de la petición HTTP.
        f_id (str): ID de la factura a eliminar.

    Returns:
        JsonResponse: Respuesta en formato JSON indicando éxito o error.
    """

    if request.method == "POST":
        try:
            factura = FacturaPdbt.objects(id=f_id).first()
            if not factura:
                return JsonResponse({"success": False, "error": "Factura no encontrada"})
            factura.delete()
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Método no permitido"})

# Exportar facturas a Excel PDBT y Triple(GDTMO, GDMTH Y GDBT)
@never_cache
@login_required_custom
def exportar_facturas_pdbt_excel_admin(request):
    """
    Exporta las facturas PDBT a un archivo Excel.
    - Filtra las facturas según los parámetros proporcionados en la URL.
    - Solo los usuarios con rol 'admin' pueden acceder a esta función.
    - Genera un archivo Excel con formato adecuado y lo envía como respuesta.
    """

    user = get_user(request)
    if not user or user.rol not in ["admin", "admin_energia", "admin_ambiental"]:
        return redirect('inicio')

    ur_id = request.GET.get('ur')
    sub_id = request.GET.get('subestacion')
    tipo_tarifa = request.GET.get('tipo_tarifa')
    anio = request.GET.get('anio')

    facturas = FacturaPdbt.objects()

    if ur_id:
        subs = Subestacion.objects(unidad_responsable=ur_id)
        facturas = facturas.filter(subestacion__in=subs)

    if sub_id:
        facturas = facturas.filter(subestacion=sub_id)

    if tipo_tarifa:
        facturas = facturas.filter(tipo_tarifa=tipo_tarifa)

    if anio:
        facturas = [f for f in facturas if f.fecha_registro.year == int(anio)]

    # Crear archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas PDBT"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    headers = [
        "Tipo de Tarifa", "Subestación", "UR", "Días de Periodo", "Periodo",
        "Consumo (kWh)", "Cargo Energía", "Importe Demanda Máxima", "DAP",
        "IVA", "Total a Pagar", "Fecha Registro", "Estado", "Creado por", "Última actualización", "Actualizado por"
    ]
    ws.append(headers)

    for col_idx, col in enumerate(ws[1], 1):
        col.font = header_font
        col.fill = header_fill
        col.alignment = center_align
        col.border = border_style

    for row_idx, factura in enumerate(facturas, start=2):
        fila = [
            factura.tipo_tarifa,
            factura.subestacion.no_servicio if factura.subestacion else '',
            factura.subestacion.unidad_responsable.nombre if factura.subestacion and factura.subestacion.unidad_responsable else '',
            factura.dias_periodo,
            factura.periodo,
            factura.consumo,
            factura.cargo_energia,
            factura.importe_demanda_maxima,
            factura.dap,
            factura.iva,
            factura.total_a_pagar,
            factura.fecha_registro.strftime('%d/%m/%Y'),
            factura.status,
            f"{factura.creado_por.nombres} {factura.creado_por.apellidos}" if factura.creado_por else '',
            factura.ultima_actualizacion.strftime('%d/%m/%Y') if factura.ultima_actualizacion else '',
            f"{factura.actualizado_por.nombres} {factura.actualizado_por.apellidos}" if factura.actualizado_por else ''
        ]
        ws.append(fila)

        for col_idx, value in enumerate(fila, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_style
            cell.alignment = center_align

            # Aplicar formato a números decimales si corresponde
            if isinstance(value, (float, int)) and col_idx in [6, 7, 8, 9, 10, 11]:
                cell.number_format = '#,##0.00'

    # Auto ajustar columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # Nombre del archivo
    nombre_ur = ''
    if ur_id:
        try:
            unidad = UnidadResponsable.objects.get(id=ur_id)
            nombre_ur = unidad.nombre.replace(" ", "_")
        except UnidadResponsable.DoesNotExist:
            nombre_ur = "UR_Desconocida"
    else:
        nombre_ur = "Todas_UR"

    anio_texto = f"_{anio}" if anio else ""
    filename = f"Facturas_PDBT_{nombre_ur}{anio_texto}.xlsx"

    # Respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    add_never_cache_headers(response)
    return response

@never_cache
@login_required_custom
def exportar_facturas_triple_excel_admin(request):
    """
    Exporta las facturas de tipo 'triple' a un archivo Excel.
    - Filtra las facturas según los parámetros proporcionados en la URL.
    - Solo los usuarios con rol 'admin' pueden acceder a esta función.
    - Genera un archivo Excel con formato adecuado y lo envía como respuesta.
    """

    user = get_user(request)
    if not user or user.rol not in ["admin", "admin_energia", "admin_ambiental"]:
        return redirect('inicio')

    ur_id = request.GET.get('ur')
    sub_id = request.GET.get('subestacion')
    tipo_tarifa = request.GET.get('tipo_tarifa')
    anio = request.GET.get('anio')

    facturas = FacturaEnergeticaTriple.objects()

    if ur_id:
        subs = Subestacion.objects(unidad_responsable=ur_id)
        facturas = facturas.filter(subestacion__in=subs)

    if sub_id:
        facturas = facturas.filter(subestacion=sub_id)

    if tipo_tarifa:
        facturas = facturas.filter(tipo_tarifa=tipo_tarifa)

    if anio:
        facturas = [f for f in facturas if f.fecha_registro.year == int(anio)]

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas Triple"

    headers = [
        "Tipo Tarifa", "No. Servicio", "Unidad Responsable", "Días Periodo", "Periodo", "Consumo (kWh)",
        "Demanda Máxima", "Factor Potencia", "Factor Carga", "Cargo Energía", "Importe Demanda Máx.",
        "Importe BT", "Importe FP", "DAP", "IVA", "Total a Pagar", "Fecha Vencimiento", "Fecha Registro",
        "Estatus", "Creado Por", "Última actualización", "Actualizado por"
    ]
    ws.append(headers)

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in facturas:
        ws.append([
            row.tipo_tarifa,
            row.subestacion.no_servicio if row.subestacion else '',
            row.subestacion.unidad_responsable.nombre if row.subestacion and row.subestacion.unidad_responsable else '',
            row.dias_periodo,
            row.periodo,
            row.consumo,
            row.demanda_maxima,
            row.factor_potencia,
            row.factor_carga,
            row.cargo_energia,
            row.importe_demanda_maxima,
            row.importe_bt,
            row.importe_fp,
            row.dap,
            row.iva,
            row.total_a_pagar,
            row.fecha_vencimiento.strftime('%d/%m/%Y') if row.fecha_vencimiento else '',
            row.fecha_registro.strftime('%d/%m/%Y'),
            row.status,
            f"{row.creado_por.nombres} {row.creado_por.apellidos}" if row.creado_por else '',
            row.ultima_actualizacion.strftime('%d/%m/%Y') if row.ultima_actualizacion else '',
            f"{row.actualizado_por.nombres} {row.actualizado_por.apellidos}" if row.actualizado_por else ''
        ])

    # Aplicar bordes y alinear celdas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajustar columnas automáticamente
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

    # Construcción del nombre del archivo
    nombre_ur = ''
    if ur_id:
        try:
            unidad = UnidadResponsable.objects.get(id=ur_id)
            nombre_ur = unidad.nombre.replace(" ", "_")
        except UnidadResponsable.DoesNotExist:
            nombre_ur = "UR_Desconocida"
    else:
        nombre_ur = "Todas_UR"

    anio_texto = f"_{anio}" if anio else ""
    filename = f"Facturas_de_{nombre_ur}{anio_texto}.xlsx"

    # Respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    add_never_cache_headers(response)
    return response