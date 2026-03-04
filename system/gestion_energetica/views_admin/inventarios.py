from bson import ObjectId

from django.contrib import messages
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views.decorators.cache import never_cache
from django.db.models import Sum

from mongoengine.errors import DoesNotExist

from system.models import InventarioClimatizacion, InventarioLuminarias, InventarioMiscelaneos, PeriodoInventario, UnidadResponsable, Subestacion, Edificio
from system.views import get_user
from system.decorators import login_required_custom

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

@never_cache
@login_required_custom
def historiales_registros_inventarios_admin(request):
    return render(request, 'systemsigo/Inventarios/historiales.html')  

@never_cache
@login_required_custom
def admin_inventarios_filtro(request):
    """
    Vista para filtrar y listar inventarios según unidad responsable, periodo y tipo.
    - Restringida a usuarios autenticados con el decorador login_required_custom.
    - never_cache: evita el almacenamiento en caché de esta vista.
    - Calcula totales de potencia, horas y consumo según el tipo seleccionado.
    - Muestra un formulario de filtro y los resultados en la misma página.
    - Si el usuario no es admin, redirige al login con un mensaje de error.
    """

    user = get_user(request)
    if not user or user.rol not in ["admin", "admin_energia", "admin_ambiental"]:
        messages.error(request, "Acceso denegado.")
        return redirect("login")

    unidades = UnidadResponsable.objects()
    periodos = PeriodoInventario.objects()
    tipos = ["Climatización", "Luminarias", "Misceláneos"]

    # Inicializar variables de totales
    total_potencia_clim = total_horas_clim = total_consumo_clim = 0
    total_potencia_lum = total_horas_lum = total_consumo_lum = 0
    total_potencia_misc = total_horas_misc = total_consumo_misc = 0

    registros = []
    unidad_id = request.GET.get("unidad")
    periodo_id = request.GET.get("periodo")
    tipo = request.GET.get("tipo")

    if unidad_id and periodo_id and tipo:
        try:
            unidad_obj = UnidadResponsable.objects.get(id=ObjectId(unidad_id))
            periodo_obj = PeriodoInventario.objects.get(id=ObjectId(periodo_id))

            modelo_map = {
                "Climatización": InventarioClimatizacion,
                "Luminarias": InventarioLuminarias,
                "Misceláneos": InventarioMiscelaneos
            }

            modelo = modelo_map.get(tipo)
            if modelo:
                registros = modelo.objects(
                    unidad_responsable=unidad_obj,
                    periodo=periodo_obj
                )

                # Calcular totales según tipo
                if tipo == "Climatización":
                    total_potencia_clim = sum([i.potencia_total or 0 for i in registros])
                    total_horas_clim = sum([i.horas_mes or 0 for i in registros])
                    total_consumo_clim = sum([i.consumo_mensual or 0 for i in registros])

                elif tipo == "Luminarias":
                    total_potencia_lum = sum([i.potencia_total_lum or 0 for i in registros])
                    total_horas_lum = sum([i.consumo_mensual_horas or 0 for i in registros])
                    total_consumo_lum = sum([i.consumo_mensual or 0 for i in registros])

                elif tipo == "Misceláneos":
                    total_potencia_misc = sum([i.potencia_total or 0 for i in registros])
                    total_horas_misc = sum([i.horas_mes or 0 for i in registros])
                    total_consumo_misc = sum([i.consumo_mensual or 0 for i in registros])

        except Exception as e:
            messages.error(request, f"Error al obtener registros: {e}")

    return render(request, "systemsigo/Inventarios/lista.html", {
        "unidades": unidades,
        "periodos": periodos,
        "tipos": tipos,
        "registros": registros,
        "unidad_seleccionada": unidad_id,
        "periodo_seleccionado": periodo_id,
        "tipo_seleccionado": tipo,
        "total_potencia_clim": total_potencia_clim,
        "total_consumo_clim": total_consumo_clim,
        "total_horas_clim": total_horas_clim,
        "total_potencia_lum": total_potencia_lum,
        "total_horas_lum": total_horas_lum,
        "total_consumo_lum": total_consumo_lum,
        "total_potencia_misc": total_potencia_misc,
        "total_horas_misc": total_horas_misc,
        "total_consumo_misc": total_consumo_misc
    })

# Inventarios - Exportar a Excel (Por tipo de inventario)
@never_cache
@login_required_custom
def exportar_excel_inventario(request):
    """
    Vista para exportar los inventarios filtrados a un archivo Excel.
    - Restringida a usuarios autenticados con el decorador login_required_custom.
    - never_cache: evita el almacenamiento en caché de esta vista.
    - Genera un archivo Excel con los datos del inventario según la unidad, periodo y tipo seleccionados.
    - Los totales de potencia, horas y consumo se calculan y se incluyen en el archivo.
    - El archivo se descarga automáticamente con un nombre basado en la unidad y periodo seleccionados.
    - Si no se encuentra la unidad o periodo, muestra un mensaje de error y redirige a la vista de filtro.
    """

    unidad_id = request.GET.get('unidad')
    periodo_id = request.GET.get('periodo')
    tipo = request.GET.get('tipo')

    try:
        unidad = UnidadResponsable.objects.get(id=unidad_id)
    except DoesNotExist:
        messages.error(request, "Unidad responsable no encontrada.")
        return redirect("admin_inventarios_filtro")

    try:
        periodo = PeriodoInventario.objects.get(id=periodo_id)
    except DoesNotExist:
        messages.error(request, "Periodo no encontrado.")
        return redirect("admin_inventarios_filtro")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{tipo}"

    total_potencia = 0
    total_consumo = 0
    total_horas_mes = 0  # si aplica

    if tipo == "Climatización":
        headers = [
            "Edificio", "Nivel", "Área", "Tipo Clima", "Marca", "Modelo", "Capacidad BTU/HR",
            "Voltaje", "Amperaje", "Potencia (Watts)", "Potencia total (kW)", "Horas al mes", "Consumo mensual (kWh/mes)",
            "Fecha de registro", "Creado por", "Actualizado por", "Última modificación"
        ]
        registros = InventarioClimatizacion.objects.filter(unidad_responsable=unidad, periodo=periodo)
        data = []
        for i in registros:
            total_potencia += i.potencia_total or 0
            total_horas_mes += i.horas_mes or 0
            total_consumo += i.consumo_mensual or 0
            data.append([
                i.edificio.nombre, i.nivel, i.area.nombre, i.tipo_clima, i.marca, i.modelo,
                i.capacidad, i.voltaje, i.amperaje, i.potencia, i.potencia_total, i.horas_mes, i.consumo_mensual,
                i.fecha_registro.strftime('%Y-%m-%d %H:%M') if i.fecha_registro else '',
                i.creado_por.nombre_completo if i.creado_por else '',
                i.actualizado_por.nombre_completo if i.actualizado_por else '',
                i.ultima_actualizacion.strftime('%Y-%m-%d %H:%M') if i.ultima_actualizacion else ''
            ])

    elif tipo == "Luminarias":
        headers = [
            "Edificio", "Área", "Nivel", "Tipo Lámpara", "N° Luminarias", "Lámparas/Luminaria",
            "Potencia por lámpara (Watts)", "Potencia Total (kW)", "Horas al mes", "Consumo mensual (kWh/mes)", "Fecha de registro",
            "Creado por", "Actualizado por", "Última modificación"
        ]
        registros = InventarioLuminarias.objects.filter(unidad_responsable=unidad, periodo=periodo)
        data = []
        for i in registros:
            total_potencia += i.potencia_total_lum or 0
            total_horas_mes += i.consumo_mensual_horas or 0
            total_consumo += i.consumo_mensual or 0
            data.append([
                i.edificio.nombre, i.area.nombre, i.nivel, i.tipo_lampara, i.num_luminarias,
                i.lamp_luminarias, i.potencia_lamp, i.potencia_total_lum, i.consumo_mensual_horas, i.consumo_mensual,
                i.fecha_registro.strftime('%Y-%m-%d %H:%M') if i.fecha_registro else '',
                f"{i.creado_por.nombre_completo} ({i.creado_por.email})" if i.creado_por else "N/A",
                f"{i.actualizado_por.nombre_completo} ({i.actualizado_por.email})" if i.actualizado_por else "N/A",
                i.ultima_actualizacion.strftime('%Y-%m-%d %H:%M') if i.ultima_actualizacion else "N/A"
            ])

    elif tipo == "Misceláneos":
        headers = [
            "Edificio", "Nivel", "Área", "Misceláneo", "Marca", "Modelo", "Voltaje", "Amperaje",
            "Potencia (Watts)", "Potencia Total (kW)", "Horas al mes", "Consumo mensual (kWh/mes)", "Fecha de registro", "Creado por", "Actualizado por", "Última modificación"
        ]
        registros = InventarioMiscelaneos.objects.filter(unidad_responsable=unidad, periodo=periodo)
        data = []
        for i in registros:
            total_potencia += i.potencia or 0
            total_horas_mes += i.horas_mes or 0
            total_consumo += i.consumo_mensual or 0
            data.append([
                i.edificio.nombre, i.nivel, i.area.nombre, i.miscelaneos, i.marca, i.modelo,
                i.voltaje, i.amperaje, i.potencia, i.potencia_total, i.horas_mes, i.consumo_mensual,
                i.fecha_registro.strftime('%Y-%m-%d %H:%M') if i.fecha_registro else '',
                f"{i.creado_por.nombre_completo} ({i.creado_por.email})" if i.creado_por else "N/A",
                f"{i.actualizado_por.nombre_completo} ({i.actualizado_por.email})" if i.actualizado_por else "N/A",
                i.ultima_actualizacion.strftime('%Y-%m-%d %H:%M') if i.ultima_actualizacion else "N/A"
            ])
    else:
        return HttpResponse("Tipo no válido", status=400)

    # Escribir cabecera con estilo
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for col_num, col in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Escribir datos
    for row in data:
        ws.append(row)

    # Escribir totales (última fila)
    total_row = [""] * len(headers)
    if tipo == "Climatización":
        total_row[10] = total_potencia  # Potencia total (Kw)
        total_row[11] = total_horas_mes
        total_row[12] = total_consumo  # Consumo mensual
    elif tipo == "Luminarias":
        total_row[7] = total_potencia
        total_row[8] = total_horas_mes
        total_row[9] = total_consumo
    elif tipo == "Misceláneos":
        total_row[8] = total_potencia
        total_row[9] = total_horas_mes
        total_row[10] = total_consumo

    ws.append(total_row)

    # Estilo para fila de totales
    total_font = Font(bold=True)
    last_row_idx = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=last_row_idx, column=col)
        cell.font = total_font
        cell.border = border

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Preparar respuesta
    anio = periodo.fecha_inicio.year
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nombre_archivo = f"Inventario_{tipo}_{unidad.nombre}_{periodo.nombre}_{anio}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

    wb.save(response)
    return response

# Inventarios - Exportar a Excel (De manera general, con todos los tipos en hojas separadas)
@never_cache
@login_required_custom
def exportar_excel_inventario_general(request):

    unidad_id = request.GET.get('unidad')
    periodo_id = request.GET.get('periodo')

    try:
        unidad = UnidadResponsable.objects.get(id=unidad_id)
        periodo = PeriodoInventario.objects.get(id=periodo_id)
    except DoesNotExist:
        messages.error(request, "Unidad o periodo no encontrado.")
        return redirect("admin_inventarios_filtro")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # eliminar hoja por defecto

    # =========================
    # CONFIGURACIÓN COMPARTIDA
    # =========================
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def aplicar_estilos(ws, headers):
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    def ajustar_columnas(ws):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

    # =========================
    # CLIMATIZACIÓN
    # =========================
    ws = wb.create_sheet("Climatización")
    headers = [
        "Edificio", "Nivel", "Área", "Tipo Clima", "Marca", "Modelo", "Capacidad BTU/HR",
        "Voltaje", "Amperaje", "Potencia (W)", "Potencia total (Kw)",
        "Horas al mes", "Consumo mensual (kWh/mes)"
    ]
    aplicar_estilos(ws, headers)

    total_potencia = total_horas = total_consumo = 0
    registros = InventarioClimatizacion.objects.filter(unidad_responsable=unidad, periodo=periodo)

    for i in registros:
        total_potencia += i.potencia_total or 0
        total_horas += i.horas_mes or 0
        total_consumo += i.consumo_mensual or 0
        ws.append([
            i.edificio.nombre, i.nivel, i.area.nombre, i.tipo_clima,
            i.marca, i.modelo, i.capacidad, i.voltaje, i.amperaje,
            i.potencia, i.potencia_total, i.horas_mes, i.consumo_mensual
        ])

    ws.append([""] * 10 + [total_potencia, total_horas, total_consumo])
    ajustar_columnas(ws)

    # =========================
    # LUMINARIAS
    # =========================
    ws = wb.create_sheet("Luminarias")
    headers = [
        "Edificio", "Área", "Nivel", "Tipo Lámpara",
        "N° Luminarias", "Lámparas/Luminaria", "Potencia (W)","Potencia Total (kW)",
        "Horas al mes", "Consumo mensual (kWh/mes)"
    ]
    aplicar_estilos(ws, headers)

    total_potencia = total_horas = total_consumo = 0
    registros = InventarioLuminarias.objects.filter(unidad_responsable=unidad, periodo=periodo)

    for i in registros:
        total_potencia += i.potencia_total_lum or 0
        total_horas += i.consumo_mensual_horas or 0
        total_consumo += i.consumo_mensual or 0
        ws.append([
            i.edificio.nombre, i.area.nombre, i.nivel, i.tipo_lampara,
            i.num_luminarias, i.lamp_luminarias, i.potencia_lamp, i.potencia_total_lum,
            i.consumo_mensual_horas, i.consumo_mensual
        ])

    ws.append([""] * 7 + [total_potencia, total_horas, total_consumo])
    ajustar_columnas(ws)

    # =========================
    # MISCELÁNEOS
    # =========================
    ws = wb.create_sheet("Misceláneos")
    headers = [
        "Edificio", "Nivel", "Área", "Tipo de misceláneo",
        "Marca", "Modelo", "Voltaje", "Amperaje", "Potencia (W)", "Potencia Total (kW)",
        "Horas al mes", "Consumo mensual"
    ]
    aplicar_estilos(ws, headers)

    total_potencia = total_horas = total_consumo = 0
    registros = InventarioMiscelaneos.objects.filter(unidad_responsable=unidad, periodo=periodo)

    for i in registros:
        total_potencia += i.potencia or 0
        total_horas += i.horas_mes or 0
        total_consumo += i.consumo_mensual or 0
        ws.append([
            i.edificio.nombre, i.nivel, i.area.nombre, i.miscelaneos,
            i.marca, i.modelo, i.voltaje, i.amperaje, i.potencia, i.potencia_total,
            i.horas_mes, i.consumo_mensual
        ])

    ws.append([""] * 9 + [total_potencia, total_horas, total_consumo])
    ajustar_columnas(ws)

    # =========================
    # RESPUESTA
    # =========================
    anio = periodo.fecha_inicio.year
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nombre_archivo = f"Inventario_General_{unidad.nombre}_{periodo.nombre}_{anio}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

    wb.save(response)
    return response

# Inventarios - Exportar a Excel (Agrupado por número de servicio, con todos los tipos en hojas separadas)
@never_cache
@login_required_custom
def exportar_excel_inventario_por_no_servicio(request):

    periodo_id = request.GET.get('periodo')

    try:
        periodo = PeriodoInventario.objects.get(id=periodo_id)
    except DoesNotExist:
        messages.error(request, "Periodo no encontrado.")
        return redirect("admin_inventarios_filtro")

    subestaciones = Subestacion.objects.all()

    mapa_servicios = {}

    for s in subestaciones:
        ur_id = str(s.unidad_responsable.id)

        if ur_id not in mapa_servicios:
            mapa_servicios[ur_id] = []

        mapa_servicios[ur_id].append(str(s.no_servicio))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def aplicar_estilos(ws, headers):
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    def ajustar_columnas(ws):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2


    unidades = UnidadResponsable.objects.all()

    for unidad in unidades:

        nombre_hoja = unidad.nombre[:30]
        ws = wb.create_sheet(nombre_hoja)

        campus = unidad.campus.nomenclatura
        servicios = mapa_servicios.get(str(unidad.id), ["Sin servicio"])
        servicio = ", ".join(servicios)

        ws.append([f"Campus: {campus}"])
        ws.append([f"Número de servicio: {servicio}"])
        ws.append([""])

        # ======= CLIMATIZACIÓN =======

        ws.append(["CLIMATIZACIÓN"])
        aplicar_estilos(ws, [
            "Edificio", "No. Servicio", "Nivel", "Área", "Tipo Clima",
            "Marca", "Modelo", "Capacidad BTU/HR", "Voltaje", "Amperaje", "Potencia (W)", "Potencia Total (kW)",
            "Horas al mes", "Consumo mensual (kWh/mes)"
        ])

        climas = InventarioClimatizacion.objects.filter(
            periodo=periodo,
            unidad_responsable=unidad
        ).select_related()

        for i in climas:

            numero_servicio = (
                str(i.edificio.subestacion.no_servicio)
                if i.edificio and i.edificio.subestacion
                else "Sin servicio"
            )

            ws.append([
                i.edificio.nombre,
                numero_servicio,
                i.nivel,
                i.area.nombre,
                i.tipo_clima,
                i.marca,
                i.modelo,
                i.capacidad,
                i.voltaje,
                i.amperaje,
                i.potencia,
                i.potencia_total,
                i.horas_mes,
                i.consumo_mensual
            ])

        ws.append([""])

        # ======= LUMINARIAS =======

        ws.append(["LUMINARIAS"])
        aplicar_estilos(ws, [
            "Edificio", "No. Servicio", "Área", "Nivel",
            "Tipo Lámpara", "N° Luminarias", "Lámparas/Luminaria", "Potencia por lámpara (W)",
            "Potencia Total (kW)", "Horas al mes",
            "Consumo mensual (kWh/mes)"
        ])

        luminarias = InventarioLuminarias.objects.filter(
            periodo=periodo,
            unidad_responsable=unidad
        ).select_related()


        for i in luminarias:

            numero_servicio = (
                str(i.edificio.subestacion.no_servicio)
                if i.edificio and i.edificio.subestacion
                else "Sin servicio"
            )

            ws.append([
                i.edificio.nombre,
                numero_servicio,
                i.area.nombre,
                i.nivel,
                i.tipo_lampara,
                i.num_luminarias,
                i.lamp_luminarias,
                i.potencia_lamp,
                i.potencia_total_lum,
                i.consumo_mensual_horas,
                i.consumo_mensual
            ])

        ws.append([""])

        # ======= MISCELÁNEOS =======

        ws.append(["MISCELÁNEOS"])
        aplicar_estilos(ws, [
            "Edificio", "No. Servicio", "Nivel", "Área",
            "Tipo de misceláneo", "Marca", "Modelo", "Voltaje", "Amperaje", "Potencia (W)", "Potencia Total (kW)",
            "Horas al mes", "Consumo mensual (kWh/mes)"
        ])

        miscelaneos = InventarioMiscelaneos.objects.filter(
            periodo=periodo,
            unidad_responsable=unidad
        ).select_related()

        for i in miscelaneos:

            numero_servicio = (
                str(i.edificio.subestacion.no_servicio)
                if i.edificio and i.edificio.subestacion
                else "Sin servicio"
            )

            ws.append([
                i.edificio.nombre,
                numero_servicio,
                i.nivel,
                i.area.nombre,
                i.miscelaneos,
                i.marca,
                i.modelo,
                i.voltaje,
                i.amperaje,
                i.potencia,
                i.potencia_total,
                i.horas_mes,
                i.consumo_mensual
            ])

        ajustar_columnas(ws)


    anio = periodo.fecha_inicio.year
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    nombre_archivo = f"Inventario_Por_Servicio{periodo.nombre}_{anio}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

    wb.save(response)
    return response

# Inventarios - Exportar a Excel (Análisis por edificio, con consumo total por tipo y mayor consumo denominado desempeño energético)
@never_cache
@login_required_custom
def exportar_analisis_consumo_edificios(request):

    unidad_id = request.GET.get('unidad')
    periodo_id = request.GET.get('periodo')

    if not unidad_id or not periodo_id:
        messages.error(request, "Debe seleccionar Unidad y Periodo.")
        return redirect("inventarios_filtro_triple")

    try:
        unidad = UnidadResponsable.objects.get(id=unidad_id)
        periodo = PeriodoInventario.objects.get(id=periodo_id)
    except DoesNotExist:
        messages.error(request, "Unidad o periodo no encontrado.")
        return redirect("inventarios_filtro_triple")

    # ===== OBTENER DATOS DE CAMPUS Y SERVICIOS =====
    subestaciones = Subestacion.objects.filter(unidad_responsable=unidad)

    numeros_servicio = ", ".join(
        str(s.no_servicio) for s in subestaciones
    ) or "Sin servicio"

    campus = unidad.campus.nomenclatura if unidad.campus else "Sin campus"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Análisis por Edificio"

    headers = [
        "Campus",
        "Número de Servicio",
        "Unidad Responsable",
        "Edificio",
        "Consumo Climatización",
        "Consumo Luminarias",
        "Consumo Misceláneos",
        "Mayor Consumo",
        "Consumo Mayor (kWh)"
    ]

    ws.append(headers)

    edificios = Edificio.objects.filter(unidad_responsable=unidad)

    for edificio in edificios:

        if edificio.subestacion and edificio.subestacion.no_servicio:
            numeros_servicio = str(edificio.subestacion.no_servicio)
        else:
            numeros_servicio = "Sin servicio"

        total_clima = sum(
            float(i.consumo_mensual or 0)
            for i in InventarioClimatizacion.objects.filter(
                unidad_responsable=unidad,
                periodo=periodo,
                edificio=edificio
            )
        )

        total_lum = sum(
            float(i.consumo_mensual or 0)
            for i in InventarioLuminarias.objects.filter(
                unidad_responsable=unidad,
                periodo=periodo,
                edificio=edificio
            )
        )

        total_misc = sum(
            float(i.consumo_mensual or 0)
            for i in InventarioMiscelaneos.objects.filter(
                unidad_responsable=unidad,
                periodo=periodo,
                edificio=edificio
            )
        )

        consumos = {
            "Climatización": total_clima,
            "Luminarias": total_lum,
            "Misceláneos": total_misc
        }

        mayor_tipo = max(consumos, key=consumos.get)
        mayor_valor = consumos[mayor_tipo]

        ws.append([
            campus,
            numeros_servicio,
            unidad.nombre,
            edificio.nombre,
            total_clima,
            total_lum,
            total_misc,
            mayor_tipo,
            mayor_valor
        ])

    # Ajustar columnas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        f'attachment; filename="Desempeño_Energético_{unidad.nombre}_{periodo.nombre}.xlsx"'
    )

    wb.save(response)
    return response