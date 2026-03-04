from django.shortcuts import render
from django.views.decorators.cache import never_cache

from django.contrib import messages
from django.shortcuts import redirect

from system.models import Edificio, UnidadResponsable, Subestacion, PeriodoInventario, InventarioClimatizacion, InventarioLuminarias, InventarioMiscelaneos
from django.http import HttpResponse
from system.decorators import login_required_custom

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

@never_cache
@login_required_custom
def listar_edificios_admin(request):
    """
    Vista para listar todos los edificios registrados en el sistema.
    - Restringida a usuarios autenticados con el decorador login_required_custom.
    - never_cache: evita el almacenamiento en caché de esta vista.
    """

    unidad_id = request.GET.get('unidad_id')
    subestacion_id = request.GET.get('subestacion_id')
    periodo_id = request.GET.get('periodo')  

    unidades = UnidadResponsable.objects.all().order_by('nombre')
    subestaciones = Subestacion.objects
    edificios = Edificio.objects
    periodos = PeriodoInventario.objects.order_by('-anio', '-mes')  

    if unidad_id:
        edificios = edificios.filter(unidad_responsable=unidad_id)
        subestaciones = subestaciones.filter(unidad_responsable=unidad_id)

    if subestacion_id:
        edificios = edificios.filter(subestacion=subestacion_id)

    return render(request, 'systemsigo/Edificios/edificios.html', {
        'edificios': edificios,
        'unidades': unidades,
        'unidad_id': unidad_id,
        'subestacion_id': subestacion_id,
        'subestaciones': subestaciones,
        'periodos': periodos,        
        'periodo_id': periodo_id,    
    })

@never_cache
@login_required_custom
def asignar_subestacion_edificio(request):
    if request.method == "POST":
        edificio_id = request.POST.get("edificio_id")
        subestacion_id = request.POST.get("subestacion_id")

        edificio = Edificio.objects(id=edificio_id).first()

        if not edificio:
            messages.error(request, "Edificio no válido.")
            return redirect("todos_edificios")


        subestacion = Subestacion.objects(
            id=subestacion_id,
            unidad_responsable=edificio.unidad_responsable
        ).first()

        if not subestacion:
            messages.error(request, "La subestación no pertenece a la misma Unidad Responsable.")
            return redirect("todos_edificios")

        edificio.subestacion = subestacion
        edificio.save()

        messages.success(request, "Número de servicio asignado correctamente.")
        return redirect("todos_edificios")

    return redirect("todos_edificios")

@never_cache
@login_required_custom
def exportar_excel_por_numero_servicio(request):

    periodo_id = request.GET.get('periodo')
    subestacion_id = request.GET.get('subestacion_id')

    if not periodo_id or not subestacion_id:
        messages.error(request, "Debe seleccionar periodo y número de servicio.")
        return redirect("todos_edificios")

    subestacion = Subestacion.objects(id=subestacion_id).first()
    periodo = PeriodoInventario.objects(id=periodo_id).first()

    if not subestacion or not periodo:
        messages.error(request, "Datos inválidos.")
        return redirect("todos_edificios")

    edificios = Edificio.objects(subestacion=subestacion)

    if not edificios:
        messages.error(request, "No hay edificios asociados a este número de servicio.")
        return redirect("todos_edificios")

    # Tomamos la UR desde el primer edificio (todos pertenecen a la misma UR)
    ur = edificios.first().unidad_responsable
    campus = ur.campus

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # =========================
    # CONFIGURACIÓN GENERAL
    # =========================
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def aplicar_estilos(ws, headers):
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=6, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    def ajustar_columnas(ws):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

    def encabezado_general(ws):
        ws["A1"] = f"Campus: {campus.nomenclatura}"
        ws["A2"] = f"Unidad Responsable: {ur.nombre}"
        ws["A3"] = f"Periodo: {periodo.nombre} | Número de Servicio: {subestacion.no_servicio}"

    # =========================
    # CLIMATIZACIÓN
    # =========================
    ws = wb.create_sheet("Climatización")
    encabezado_general(ws)

    headers = [
        "Edificio", "Nivel", "Área", "Tipo Clima", "Marca", "Modelo",
        "Capacidad BTU/HR", "Voltaje", "Amperaje",
        "Potencia (W)", "Potencia Total (Kw)",
        "Horas al mes", "Consumo mensual (kWh/mes)"
    ]

    ws.append([])
    ws.append([])
    aplicar_estilos(ws, headers)

    total_consumo = 0
    fila_inicio = ws.max_row + 1

    total_potencia_kw = 0
    total_horas = 0
    total_consumo = 0

    for edificio in edificios:
        registros = InventarioClimatizacion.objects(
            edificio=edificio,
            periodo=periodo
        )

        for i in registros:
            total_potencia_kw += i.potencia_total or 0
            total_horas += i.horas_mes or 0
            total_consumo += i.consumo_mensual or 0

            ws.append([
                edificio.nombre,
                i.nivel,
                i.area.nombre,
                i.tipo_clima,
                i.marca,
                i.modelo,
                i.capacidad,
                i.voltaje,
                i.amperaje,
                i.potencia,
                i.potencia_total,
                i.horas_mes,
                i.consumo_mensual
            ])

    ws.append([""] * 9 + ["TOTAL:",
                          total_potencia_kw,
                          total_horas,
                          total_consumo])
    
    # =========================
    # LUMINARIAS
    # =========================
    ws = wb.create_sheet("Luminarias")
    encabezado_general(ws)

    headers = [
        "Edificio", "Área", "Nivel", "Tipo Lámpara",
        "N° Luminarias", "Lámpara/Luminaria", "Potencia (W)", "Potencia Total (kW)",
        "Horas al mes", "Consumo mensual (kWh/mes)"
    ]

    ws.append([])
    ws.append([])
    aplicar_estilos(ws, headers)

    total_consumo = 0

    total_potencia_kw = 0
    total_horas = 0
    total_consumo = 0

    for edificio in edificios:
        registros = InventarioLuminarias.objects(
            edificio=edificio,
            periodo=periodo
        )

        for i in registros:
            total_potencia_kw += i.potencia_total_lum or 0
            total_horas += i.consumo_mensual_horas or 0
            total_consumo += i.consumo_mensual or 0

            ws.append([
                edificio.nombre,
                i.area.nombre,
                i.nivel,
                i.tipo_lampara,
                i.num_luminarias,
                i.lamp_luminarias,
                i.potencia_lamp,
                i.potencia_total_lum,
                i.consumo_mensual_horas,
                i.consumo_mensual
            ])

    ws.append([""] * 6 + ["TOTAL:",
                          total_potencia_kw,
                          total_horas,
                          total_consumo])

    # =========================
    # MISCELÁNEOS
    # =========================
    ws = wb.create_sheet("Misceláneos")
    encabezado_general(ws)

    headers = [
        "Edificio", "Nivel", "Área", "Tipo de misceláneo",
        "Marca", "Modelo", "Voltaje", "Amperaje", "Potencia (W)", "Potencia Total (kW)",
        "Horas al mes", "Consumo mensual (kWh/mes)"
    ]

    ws.append([])
    ws.append([])
    aplicar_estilos(ws, headers)

    total_consumo = 0

    for edificio in edificios:
        registros = InventarioMiscelaneos.objects(
            edificio=edificio,
            periodo=periodo
        )

    total_potencia = 0
    total_horas = 0
    total_consumo = 0

    for edificio in edificios:
        registros = InventarioMiscelaneos.objects(
            edificio=edificio,
            periodo=periodo
        )

        for i in registros:
            total_potencia += i.potencia or 0
            total_horas += i.horas_mes or 0
            total_consumo += i.consumo_mensual or 0

            ws.append([
                edificio.nombre,
                i.nivel,
                i.area.nombre,
                i.miscelaneos,
                i.marca,
                i.modelo,
                i.voltaje,
                i.amperaje,
                i.potencia,
                i.potencia_total,
                i.horas_mes,
                i.consumo_mensual
            ])

    ws.append([""] * 8 + ["TOTAL:",
                          total_potencia,
                          total_horas,
                          total_consumo])

    # =========================
    # RESPUESTA
    # =========================
    anio = periodo.fecha_inicio.year
    nombre_archivo = f"{ur.nombre}-{campus.nomenclatura}-{periodo.nombre}-{subestacion.no_servicio}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

    wb.save(response)
    return response